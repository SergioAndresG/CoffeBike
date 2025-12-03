from fastapi import APIRouter, Response, Depends
from fastapi.responses import FileResponse
from openpyxl import Workbook
from io import BytesIO
from sqlalchemy.orm import Session
from app.conexion import get_db
from app.models.productos import Productos
from app.models.detalle_factura import DetalleFactura
from app.models.materia_prima_recetas import MateriaPrimaRecetas
from app.models.materia_prima import MateriaPrima
from app.models.usuario import Usuarios
from app.models.factura import Factura
import os
from datetime import datetime
from app.schemas.alerta_schemas import EmailRequest
from .productos import send_email_task
import requests


router_reporte = APIRouter()

from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

@router_reporte.get("/exportar_excel/")
def exportar_excel(db: Session = Depends(get_db)):
    detalles = db.query(DetalleFactura).all()
    wb = Workbook()

    def auto_ajustar_columnas(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # 1. Materias Primas
    ws_mp = wb.active
    ws_mp.title = "Materias Primas"
    ws_mp.append(["Nombre", "Cantidad en inventario", "Precio unitario", "Costo total"])
    materias_primas = db.query(MateriaPrima).all()
    for mp in materias_primas:
        costo_total = float(mp.cantidad) * float(mp.precio_unitario)
        ws_mp.append([mp.nombre, mp.cantidad, mp.precio_unitario, costo_total])
    auto_ajustar_columnas(ws_mp)

    # 2. Productos Preparados
    ws_pp = wb.create_sheet("Productos Preparados")
    ws_pp.append(["Producto", "Ingredientes usados", "Costo producción", "Precio venta", "Margen ganancia"])
    productos_preparados = db.query(Productos).filter(Productos.tipo == "HECHO").all()
    for p in productos_preparados:
        ingredientes = db.query(MateriaPrimaRecetas).filter(MateriaPrimaRecetas.producto_id == p.id).all()
        nombres = []
        costo_total = 0.0
        for ing in ingredientes:
            mp = db.query(MateriaPrima).filter(MateriaPrima.id == ing.materia_prima_id).first()
            nombres.append(mp.nombre)
            costo_total += float(ing.cantidad_ingrediente) * float(mp.precio_unitario)
        margen = float(p.precio_salida or 0) - costo_total
        ws_pp.append([p.nombre, ", ".join(nombres), costo_total, p.precio_salida, margen])
    auto_ajustar_columnas(ws_pp)

    # 3. Productos Comprados
    ws_pc = wb.create_sheet("Productos Comprados")
    ws_pc.append(["Producto", "Precio entrada", "Precio venta", "Margen"])
    productos_comprados = db.query(Productos).filter(Productos.tipo == "COMPRADO").all()
    for p in productos_comprados:
        margen = float(p.precio_salida or 0) - float(p.precio_entrada or 0)
        ws_pc.append([p.nombre, p.precio_entrada, p.precio_salida, margen])
    auto_ajustar_columnas(ws_pc)

    # 4. Ventas
    ws_v = wb.create_sheet("Ventas")
    ws_v.append(["Fecha", "Producto", "Cantidad", "Precio unitario", "Total por producto", "ID Factura"])
    facturas = db.query(Factura).all()
    for f in facturas:
        for detalle in db.query(DetalleFactura).filter(DetalleFactura.factura_id == f.id).all():
            total = float(detalle.cantidad) * float(detalle.precio_unitario)
            row = [
                f.fecha_compra,
                detalle.producto.nombre if detalle.producto else "Producto eliminado",
                detalle.cantidad,
                detalle.precio_unitario,
                total,
                f.id
            ]
            ws_v.append(row)
            ws_v.cell(row=ws_v.max_row, column=1).number_format = "DD/MM/YYYY"
    auto_ajustar_columnas(ws_v)

    # 5. Reporte Diario
    ws5 = wb.create_sheet("Reporte Diario")
    ws5.append(["Fecha", "Total del día", "Producto más vendido"])
    ventas_por_dia = {}
    conteo_productos = {}
    for d in detalles:
        fecha = d.factura.fecha_compra.date()
        ventas_por_dia.setdefault(fecha, 0)
        ventas_por_dia[fecha] += float(d.cantidad) * float(d.precio_unitario)
        conteo_productos.setdefault(fecha, {})
        conteo_productos[fecha].setdefault(d.producto.nombre, 0)
        conteo_productos[fecha][d.producto.nombre] += d.cantidad

    fechas_ordenadas = sorted(ventas_por_dia.keys())
    for fecha in fechas_ordenadas:
        total = ventas_por_dia[fecha]
        mas_vendido = max(conteo_productos[fecha], key=conteo_productos[fecha].get)
        ws5.append([fecha, round(total, 2), mas_vendido])
        ws5.cell(row=ws5.max_row, column=1).number_format = "DD/MM/YYYY"
    auto_ajustar_columnas(ws5)

    # 6. Costos y Ganancias
    ws6 = wb.create_sheet("Costos y Ganancias")
    ws6.append(["Fecha", "Costo MP usada", "Ingresos por ventas", "Ganancia bruta", "% Margen"])
    for fecha in fechas_ordenadas:
        ingresos = ventas_por_dia[fecha]
        detalles_fecha = [d for d in detalles if d.factura.fecha_compra.date() == fecha]
        costo_usado = 0
        for d in detalles_fecha:
            if d.producto.tipo == "HECHO":
                costo_usado += sum([float(r.ingrediente.precio_unitario) * float(r.cantidad) for r in d.producto.receta]) * d.cantidad
            else:
                costo_usado += float(d.producto.precio_entrada) * d.cantidad
        ganancia = ingresos - costo_usado
        margen_pct = (ganancia / ingresos * 100) if ingresos != 0 else 0
        ws6.append([fecha, round(costo_usado, 2), round(ingresos, 2), round(ganancia, 2), round(margen_pct, 2)])
        ws6.cell(row=ws6.max_row, column=1).number_format = "DD/MM/YYYY"
    auto_ajustar_columnas(ws6)

    # 7. Ranking
    ws7 = wb.create_sheet("Ranking Productos")
    ws7.append(["Producto", "Cantidad vendida", "Total ingresos", "Margen total"])
    ranking = {}
    for d in detalles:
        nombre = d.producto.nombre
        if nombre not in ranking:
            ranking[nombre] = {"cantidad": 0, "ingresos": 0, "margen": 0, "tipo": d.producto.tipo}
        ranking[nombre]["cantidad"] += d.cantidad
        ranking[nombre]["ingresos"] += float(d.precio_unitario) * d.cantidad
        if d.producto.tipo == "HECHO":
            costo = sum([float(r.ingrediente.precio_unitario) * float(r.cantidad) for r in d.producto.receta])
        else:
            costo = float(d.producto.precio_entrada)
        ranking[nombre]["margen"] += (float(d.precio_unitario) - costo) * d.cantidad

    ranking_ordenado = sorted(ranking.items(), key=lambda x: x[1]["cantidad"], reverse=True)
    for nombre, datos in ranking_ordenado:
        ws7.append([nombre, datos["cantidad"], round(datos["ingresos"], 2), round(datos["margen"], 2)])
    auto_ajustar_columnas(ws7)

    # Guardar archivo
    filename = f"reporte_inventario_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.abspath(os.path.join("reportes", filename))
    print(f"Ruta absoluta del archivo: {filepath}")
    os.makedirs("reportes", exist_ok=True)
    wb.save(filepath)


    usuario = db.query(Usuarios).filter(Usuarios.rol == "JEFE").first()
    if usuario and usuario.correo:
            asunto = "¡El Reporte Semanal ha llegado!"
            mensaje = "Adjunto se encuntra el reporte generado"
            try:
                email_request = {
                    "to_email":usuario.correo,
                    "subject":asunto,
                    "message":mensaje
                }
                
                response = requests.post("http://127.0.0.1:8000/productos/send-email", json=email_request, params={"attachment_path": filepath})

                if response.status_code == 200:
                    print(f"Alerta enviada a {usuario.nombre}")
                else:
                    print(f"Error al enviar la alerta del stock")
            except Exception as e:
                print(f"Error al enviar el mensaje: {e}")



    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")